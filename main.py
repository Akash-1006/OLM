# main.py  ── FastAPI entry point  (replaces app.py)
# Run with:  uvicorn main:app --host 0.0.0.0 --port 8000
# ─────────────────────────────────────────────────────────────────────────────
import asyncio
import csv
import io
import json
import os
import threading
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import (
    Depends, FastAPI, File, Form, Header, HTTPException,
    Query, Request, UploadFile,
)
from fastapi.responses import HTMLResponse, JSONResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update, WebAppInfo
from telegram.ext import Application, CallbackQueryHandler, CommandHandler

from bot.daily_digest import register_daily_digest
from bot.handlers.admin import register_admin_handlers
from bot.handlers.followup import handle_followup_response
from bot.handlers.lead import start_lead
from bot.scheduler import (
    print_scheduler_status,
    reschedule_on_update,
    schedule_followups,
    scheduler as apscheduler,
    sync_scheduler_with_db,
)
from config import TELEGRAM_TOKEN, WEBHOOK_URL
from db import Base, SessionLocal, engine
from models.exec_target import ExecTarget
from models.lead import Lead
from models.lead_update import LeadUpdate

# ══════════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════
IST_OFFSET = timedelta(hours=5, minutes=30)

ADMIN_PASSWORD: str = os.getenv("ADMIN_PASSWORD", "mcube@admin123")

# ══════════════════════════════════════════════════════════════════════════════
#  APP INIT
# ══════════════════════════════════════════════════════════════════════════════
Base.metadata.create_all(engine)

app = FastAPI(title="OLM API", version="2.0.0")

# Serve static files (CSS/JS/images used by admin & miniapp)
app.mount("/static", StaticFiles(directory="static"), name="static")

# ══════════════════════════════════════════════════════════════════════════════
#  TELEGRAM BOT SETUP  (identical logic to Flask version)
# ══════════════════════════════════════════════════════════════════════════════
telegram_app = Application.builder().token(TELEGRAM_TOKEN).build()

telegram_app.add_handler(CommandHandler("newlead", start_lead))
telegram_app.add_handler(CommandHandler("start",   start_lead))
telegram_app.add_handler(CallbackQueryHandler(handle_followup_response, pattern="^fu_"))
register_admin_handlers(telegram_app)

def _start_loop(loop: asyncio.AbstractEventLoop) -> None:
    asyncio.set_event_loop(loop)
    loop.run_forever()

bot_loop = asyncio.new_event_loop()
threading.Thread(target=_start_loop, args=(bot_loop,), daemon=True).start()

async def _init_bot() -> None:
    await telegram_app.initialize()
    await telegram_app.start()
    await telegram_app.bot.set_webhook(f"{WEBHOOK_URL}/webhook")
    print("✅ Telegram bot initialized & webhook set")

asyncio.run_coroutine_threadsafe(_init_bot(), bot_loop).result(timeout=15)

register_daily_digest(apscheduler, telegram_app.bot, bot_loop)
sync_scheduler_with_db(telegram_app.bot, bot_loop)

apscheduler.add_job(
    sync_scheduler_with_db,
    trigger="interval",
    hours=1,
    args=[telegram_app.bot, bot_loop],
    id="sync_scheduler",
    replace_existing=True,
)
print_scheduler_status()

# ══════════════════════════════════════════════════════════════════════════════
#  SECURITY DEPENDENCIES
# ══════════════════════════════════════════════════════════════════════════════

def require_admin(
    x_admin_key: Optional[str] = Header(None),
    key: Optional[str] = Query(None),
) -> None:
    """Protect admin endpoints. Accepts X-Admin-Key header OR ?key= query param."""
    provided = x_admin_key or key or ""
    if provided != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Unauthorized")


def require_app_auth(
    request: Request,
    x_access_key: Optional[str] = Header(None),
    key: Optional[str] = Query(None),
) -> None:
    """Protect miniapp endpoints. Only enforced when MINIAPP_ACCESS_KEY is set."""
    access_key = os.getenv("MINIAPP_ACCESS_KEY", "").strip()
    if not access_key:
        return  # open access (legacy)
    provided = x_access_key or key or ""
    if provided != access_key:
        raise HTTPException(status_code=401, detail="Unauthorized")


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _to_ist_str(dt: Optional[datetime]) -> str:
    if not dt:
        return ""
    return (dt + IST_OFFSET).strftime("%d %b %Y %I:%M %p")


def _is_overdue(lead: Lead) -> bool:
    if lead.site_status in ("Won", "Lost") or not lead.last_followup_at:
        return False
    if lead.last_user_update_at and lead.last_user_update_at >= lead.last_followup_at:
        return False
    IST = timezone(IST_OFFSET)
    lfa_ist  = lead.last_followup_at.replace(tzinfo=timezone.utc).astimezone(IST)
    deadline = lfa_ist.replace(hour=18, minute=0, second=0, microsecond=0)
    return datetime.now(timezone.utc).astimezone(IST) > deadline


def _inject_base_url(html: str) -> str:
    base_url = (os.getenv("MINIAPP_BASE_URL") or os.getenv("WEBHOOK_URL", "")).rstrip("/")
    return html.replace("https://YOUR_NGROK_OR_DOMAIN", base_url)


def _read_html(path: str) -> str:
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def _safe_float(v) -> Optional[float]:
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


# ══════════════════════════════════════════════════════════════════════════════
#  TELEGRAM WEBHOOK
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/webhook")
async def webhook(request: Request) -> Response:
    data   = await request.json()
    update = Update.de_json(data, telegram_app.bot)
    asyncio.run_coroutine_threadsafe(telegram_app.process_update(update), bot_loop)
    return Response(content="OK")


# ══════════════════════════════════════════════════════════════════════════════
#  MINIAPP — SERVE HTML
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/miniapp", response_class=HTMLResponse)
def serve_miniapp() -> HTMLResponse:
    html = _inject_base_url(_read_html(os.path.join("miniapp", "index.html")))
    return HTMLResponse(content=html, headers={"Cache-Control": "no-store, no-cache, must-revalidate"})


@app.get("/miniapp/update", response_class=HTMLResponse)
def serve_miniapp_update() -> HTMLResponse:
    """Same HTML — JS switches to update mode when ?lead_id= is present."""
    html = _inject_base_url(_read_html(os.path.join("miniapp", "index.html")))
    return HTMLResponse(content=html, headers={"Cache-Control": "no-store, no-cache, must-revalidate"})


# ══════════════════════════════════════════════════════════════════════════════
#  MINIAPP — SUBMIT LEAD
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/submit_lead")
async def submit_lead(
    _auth: None = Depends(require_app_auth),
    company_name:         str = Form(...),
    contact_name:         str = Form(...),
    phone:                str = Form(...),
    work_status:          str = Form(...),
    stage:                str = Form(...),
    material:             str = Form(...),
    quantity:             str = Form(...),
    next_followup_date:   str = Form(...),
    grade:                str = Form(""),
    remarks:              str = Form(""),
    latitude:             Optional[str] = Form(None),
    longitude:            Optional[str] = Form(None),
    tg_user:              str = Form("{}"),
    photos:               list[UploadFile] = File(default=[]),
) -> JSONResponse:
    # ── Parse IST → UTC for next_followup_date ────────────────────────────────
    followup_dt: Optional[datetime] = None
    if next_followup_date:
        try:
            followup_dt = datetime.fromisoformat(next_followup_date) - IST_OFFSET
        except ValueError:
            raise HTTPException(400, "Invalid next_followup_date format. Expected YYYY-MM-DDTHH:MM.")

    # ── Location ──────────────────────────────────────────────────────────────
    try:
        lat = float(latitude)  if latitude  else None
        lng = float(longitude) if longitude else None
    except (ValueError, TypeError):
        lat = lng = None

    if not lat or not lng:
        raise HTTPException(400, "Location is required")

    # ── Telegram user ─────────────────────────────────────────────────────────
    try:
        tg = json.loads(tg_user)
    except (ValueError, TypeError):
        tg = {}
    chat_id   = tg.get("id")
    user_name = (
        tg.get("name")
        or " ".join(filter(None, [tg.get("first_name", ""), tg.get("last_name", "")]))
        or tg.get("username")
        or "Unknown"
    )

    # ── Save photos ───────────────────────────────────────────────────────────
    UPLOAD_DIR = os.path.join("uploads", "leads")
    os.makedirs(UPLOAD_DIR, exist_ok=True)

    photo_paths: list[str] = []
    for upload in photos:
        if upload and upload.filename:
            safe_name = f"{chat_id or 'unknown'}_{upload.filename}"
            save_path = os.path.join(UPLOAD_DIR, safe_name)
            content   = await upload.read()
            with open(save_path, "wb") as f:
                f.write(content)
            photo_paths.append(save_path)

    if not photo_paths:
        raise HTTPException(400, "At least one photo is required")

    # ── Persist to DB ─────────────────────────────────────────────────────────
    session = SessionLocal()
    try:
        lead = Lead(
            company_name        = company_name.strip(),
            client_name         = contact_name.strip(),
            client_phone        = phone.strip(),
            site_status         = work_status.strip(),
            stage               = stage.strip(),
            material            = material.strip(),
            grade               = grade.strip(),
            quantity            = quantity.strip(),
            remarks             = remarks.strip(),
            photo_paths         = ",".join(photo_paths),
            latitude            = lat,
            longitude           = lng,
            location            = f"{lat:.5f},{lng:.5f}",
            sales_exec_id       = chat_id,
            sales_exec_name     = user_name,
            next_followup_date  = followup_dt,
            last_user_update_at = datetime.utcnow(),
        )
        session.add(lead)
        session.commit()
        session.refresh(lead)

        if chat_id:
            schedule_followups(
                lead_id = lead.id,
                chat_id = chat_id,
                bot     = telegram_app.bot,
                loop    = bot_loop,
                stage   = stage,
            )

        print(f"✅ Lead #{lead.id} saved — {company_name} | {stage} | {material}")
        return JSONResponse({"status": "ok", "lead_id": lead.id})

    except Exception as exc:
        session.rollback()
        print(f"❌ DB error: {exc}")
        raise HTTPException(500, "Database error")
    finally:
        session.close()


# ══════════════════════════════════════════════════════════════════════════════
#  MINIAPP — GENERATE QUOTE
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/generate_quote")
async def generate_quote(
    _auth: None = Depends(require_app_auth),
    request: Request = None,
) -> JSONResponse:
    from bot.quote_generator import build_quote_pdf

    data      = await request.json()
    company   = (data.get("company_name") or "").strip()
    quantity  = (data.get("quantity")     or "").strip()
    grade     = (data.get("grade")        or "").strip()
    rate      = (data.get("rate")         or "").strip()
    location  = (data.get("location")     or "").strip()
    user_id   = data.get("user_id")

    try:
        tg = json.loads(data["tg_user"]) if isinstance(data.get("tg_user"), str) else data.get("tg_user", {})
    except Exception:
        tg = {}

    chat_id   = tg.get("id") or user_id
    exec_name = (
        " ".join(filter(None, [tg.get("first_name", ""), tg.get("last_name", "")])) or
        tg.get("username") or "Executive"
    ).strip()

    if not all([company, quantity, grade, rate, location, chat_id]):
        raise HTTPException(400, "Missing required fields")

    TEMPLATE_PATH = os.path.join(os.getenv("QUOTE_TEMPLATE_DIR", "static"), "quote_template.pdf")
    if not os.path.exists(TEMPLATE_PATH):
        raise HTTPException(500, f"Quote template not found at '{TEMPLATE_PATH}'.")

    buf = io.BytesIO()
    try:
        build_quote_pdf(
            out_buf=buf, template_path=TEMPLATE_PATH,
            company=company, location=location,
            quantity=quantity, grade=grade, rate=rate, exec_name=exec_name,
        )
        buf.seek(0)
    except Exception as exc:
        raise HTTPException(500, f"PDF generation failed: {exc}")

    now_ist   = datetime.now(timezone(IST_OFFSET))
    date_str  = now_ist.strftime("%d %b %Y")
    filename  = f"Titans_Quote_{company.replace(' ', '_')}_{date_str.replace(' ', '')}.pdf"

    try:
        qty_f  = float(quantity)
    except (ValueError, TypeError):
        qty_f  = 0.0
    try:
        rate_f = float(rate)
    except (ValueError, TypeError):
        rate_f = 0.0

    async def _send_pdf() -> None:
        await telegram_app.bot.send_document(
            chat_id   = int(chat_id),
            document  = buf,
            filename  = filename,
            caption   = (
                f"📄 *Quote — {company}*\n"
                f"📍 {location}  |  🏗 {grade}  |  📦 {qty_f:.0f} cum\n"
                f"💰 Rate: ₹{rate_f:,.0f}/cum (incl. GST)"
            ),
            parse_mode = "Markdown",
        )

    future = asyncio.run_coroutine_threadsafe(_send_pdf(), bot_loop)
    try:
        future.result(timeout=30)
        return JSONResponse({"status": "ok", "filename": filename})
    except Exception as exc:
        raise HTTPException(500, str(exc))


# ══════════════════════════════════════════════════════════════════════════════
#  MINIAPP — AUTH / LEAD DETAIL / UPDATE / HISTORY / GOALS
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/api/auth")
async def api_auth(request: Request) -> JSONResponse:
    data       = await request.json()
    key        = (data.get("key") or "").strip()
    access_key = os.getenv("MINIAPP_ACCESS_KEY", "").strip()
    if not access_key or key == access_key:
        return JSONResponse({"status": "ok"})
    raise HTTPException(401, "Invalid access key")


@app.get("/api/lead/{lead_id}")
def api_lead_detail(
    lead_id: int,
    user_id: int = Query(...),
    _auth: None = Depends(require_app_auth),
) -> JSONResponse:
    session = SessionLocal()
    try:
        lead = (session.query(Lead)
                .filter(Lead.id == lead_id, Lead.sales_exec_id == user_id)
                .first())
        if not lead:
            raise HTTPException(404, "Lead not found")

        nfd_ist = None
        if lead.next_followup_date:
            nfd_aware = lead.next_followup_date.replace(tzinfo=timezone.utc)
            nfd_ist   = (nfd_aware + IST_OFFSET).strftime("%Y-%m-%dT%H:%M")

        return JSONResponse({
            "id":           lead.id,
            "company_name": lead.company_name,
            "client_name":  lead.client_name,
            "client_phone": lead.client_phone,
            "site_status":  lead.site_status,
            "stage":        lead.stage,
            "material":     lead.material,
            "grade":        lead.grade,
            "quantity":     lead.quantity,
            "remarks":      lead.remarks,
            "latitude":     lead.latitude,
            "longitude":    lead.longitude,
            "next_followup_date": nfd_ist,
        })
    finally:
        session.close()


@app.post("/api/lead/{lead_id}/update")
async def api_update_lead(
    lead_id: int,
    _auth: None = Depends(require_app_auth),
    request: Request = None,
    company_name:        Optional[str] = Form(None),
    contact_name:        Optional[str] = Form(None),
    phone:               Optional[str] = Form(None),
    work_status:         Optional[str] = Form(None),
    stage:               Optional[str] = Form(None),
    material:            Optional[str] = Form(None),
    grade:               Optional[str] = Form(None),
    quantity:            Optional[str] = Form(None),
    remarks:             Optional[str] = Form(None),
    next_followup_date:  Optional[str] = Form(None),
    latitude:            Optional[str] = Form(None),
    longitude:           Optional[str] = Form(None),
    tg_user:             str = Form("{}"),
    photos:              list[UploadFile] = File(default=[]),
) -> JSONResponse:
    try:
        tg = json.loads(tg_user)
    except Exception:
        tg = {}
    user_id   = tg.get("id")
    user_name = (
        tg.get("name")
        or " ".join(filter(None, [tg.get("first_name", ""), tg.get("last_name", "")]))
        or tg.get("username") or "Unknown"
    )
    if not user_id:
        raise HTTPException(400, "user_id is required")

    session = SessionLocal()
    try:
        lead = (session.query(Lead)
                .filter(Lead.id == lead_id, Lead.sales_exec_id == user_id)
                .first())
        if not lead:
            raise HTTPException(404, "Lead not found")

        if company_name: lead.company_name = company_name.strip()
        if contact_name: lead.client_name  = contact_name.strip()
        if phone:        lead.client_phone = phone.strip()
        if work_status:  lead.site_status  = work_status.strip()
        if stage:        lead.stage        = stage.strip()
        if material:     lead.material     = material.strip()
        if grade:        lead.grade        = grade.strip()
        if quantity:     lead.quantity     = quantity.strip()
        if remarks:      lead.remarks      = remarks.strip()

        if next_followup_date:
            try:
                lead.next_followup_date = datetime.fromisoformat(next_followup_date) - IST_OFFSET
            except ValueError:
                raise HTTPException(400, "Invalid next_followup_date format.")

        try:
            if latitude:  lead.latitude  = float(latitude)
            if longitude: lead.longitude = float(longitude)
            if latitude and longitude:
                lead.location = f"{lead.latitude:.5f},{lead.longitude:.5f}"
        except (ValueError, TypeError):
            pass

        if photos:
            UPLOAD_DIR = os.path.join("uploads", "leads")
            os.makedirs(UPLOAD_DIR, exist_ok=True)
            existing = lead.photo_paths.split(",") if lead.photo_paths else []
            for upload in photos:
                if upload and upload.filename:
                    safe_name = f"{user_id}_{upload.filename}"
                    save_path = os.path.join(UPLOAD_DIR, safe_name)
                    content   = await upload.read()
                    with open(save_path, "wb") as f:
                        f.write(content)
                    existing.append(save_path)
            lead.photo_paths = ",".join(existing)

        lead.last_user_update_at = datetime.utcnow()

        snapshot = LeadUpdate(
            lead_id         = lead.id,
            sales_exec_id   = lead.sales_exec_id,
            sales_exec_name = user_name,
            company_name    = lead.company_name,
            client_name     = lead.client_name,
            client_phone    = lead.client_phone,
            site_status     = lead.site_status,
            stage           = lead.stage,
            material        = lead.material,
            grade           = lead.grade,
            quantity        = lead.quantity,
            remarks         = lead.remarks,
        )
        session.add(snapshot)
        session.commit()

        reschedule_on_update(lead.id, lead.sales_exec_id, lead.stage, telegram_app.bot, bot_loop)

        return JSONResponse({"status": "ok", "lead_id": lead.id})
    except HTTPException:
        raise
    except Exception as exc:
        session.rollback()
        raise HTTPException(500, str(exc))
    finally:
        session.close()


@app.get("/api/goals")
def api_goals(
    user_id: int = Query(...),
    _auth: None = Depends(require_app_auth),
) -> JSONResponse:
    session = SessionLocal()
    try:
        now   = datetime.now(timezone.utc)
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        q     = session.query(Lead).filter(Lead.sales_exec_id == user_id)
        qm    = q.filter(Lead.created_at >= start)

        this_month = qm.count()
        won        = qm.filter(Lead.site_status == "Won").count()
        total_qty  = sum(
            float(l.quantity) for l in qm.filter(Lead.quantity.isnot(None)).all()
            if l.quantity and _safe_float(l.quantity) is not None
        )

        target = session.query(ExecTarget).filter(ExecTarget.sales_exec_id == user_id).first()
        return JSONResponse({
            "this_month":     this_month,
            "won":            won,
            "total_quantity": round(total_qty, 2),
            "goals": {
                "monthly_target":    target.monthly_leads  if target else 30,
                "conversion_target": target.conversion_pct if target else 40,
                "volume_target":     target.volume_m3      if target else 500,
            },
        })
    finally:
        session.close()


@app.get("/api/history")
def api_history(
    user_id: int = Query(...),
    _auth: None = Depends(require_app_auth),
) -> JSONResponse:
    session = SessionLocal()
    try:
        leads = (session.query(Lead)
                 .filter(Lead.sales_exec_id == user_id)
                 .order_by(Lead.created_at.desc())
                 .limit(50).all())

        lead_ids = [l.id for l in leads]
        updates  = (session.query(LeadUpdate)
                    .filter(LeadUpdate.lead_id.in_(lead_ids))
                    .order_by(LeadUpdate.updated_at.asc())
                    .all()) if lead_ids else []

        updates_by_lead: dict = {}
        for u in updates:
            updates_by_lead.setdefault(u.lead_id, []).append({
                "id":          u.id,
                "site_status": u.site_status,
                "stage":       u.stage,
                "material":    u.material,
                "quantity":    u.quantity,
                "remarks":     u.remarks,
                "updated_at":  u.updated_at.isoformat() if u.updated_at else None,
            })

        return JSONResponse({"leads": [{
            "id":                 l.id,
            "company_name":       l.company_name,
            "client_name":        l.client_name,
            "client_phone":       l.client_phone,
            "site_status":        l.site_status,
            "stage":              l.stage,
            "material":           l.material,
            "grade":              l.grade,
            "quantity":           l.quantity,
            "remarks":            l.remarks,
            "location":           l.location,
            "created_at":         l.created_at.isoformat()         if l.created_at         else None,
            "last_followup_at":   l.last_followup_at.isoformat()   if l.last_followup_at   else None,
            "next_followup_date": l.next_followup_date.isoformat() if l.next_followup_date else None,
            "is_overdue":         _is_overdue(l),
            "updates":            updates_by_lead.get(l.id, []),
        } for l in leads]})
    finally:
        session.close()


# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN — SERVE DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════

def _serve_dashboard() -> HTMLResponse:
    base_url = (os.getenv("MINIAPP_BASE_URL") or os.getenv("WEBHOOK_URL", "")).rstrip("/")
    html     = _read_html(os.path.join("admin", "dashboard.html"))
    html     = html.replace("ADMIN_BASE_URL_PLACEHOLDER", base_url)
    return HTMLResponse(content=html, headers={"Cache-Control": "no-store, no-cache, must-revalidate"})


@app.get("/",      response_class=HTMLResponse)
@app.get("/admin", response_class=HTMLResponse)
def admin_dashboard() -> HTMLResponse:
    return _serve_dashboard()


# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN API — OVERVIEW / LEADS / LEAD DETAIL / LEAD UPDATE
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/admin/api/overview")
def admin_overview(_auth: None = Depends(require_admin)) -> JSONResponse:
    from sqlalchemy import func
    session = SessionLocal()
    try:
        total       = session.query(Lead).count()
        won         = session.query(Lead).filter(Lead.site_status == "Won").count()
        lost        = session.query(Lead).filter(Lead.site_status == "Lost").count()
        in_progress = max(total - won - lost, 0)

        candidates = session.query(Lead).filter(
            Lead.site_status.notin_(["Won", "Lost"]),
            Lead.last_followup_at.isnot(None),
        ).all()
        overdue_count = sum(1 for l in candidates if _is_overdue(l))

        recent   = session.query(Lead).order_by(Lead.created_at.desc()).limit(15).all()
        lead_ids = [l.id for l in recent]
        latest_updates = dict(
            session.query(LeadUpdate.lead_id, func.max(LeadUpdate.updated_at))
            .filter(LeadUpdate.lead_id.in_(lead_ids))
            .group_by(LeadUpdate.lead_id).all()
        ) if lead_ids else {}

        return JSONResponse({
            "total": total, "won": won, "lost": lost,
            "in_progress": in_progress, "overdue": overdue_count,
            "recent": [{
                "id":              l.id,
                "company_name":    l.company_name,
                "client_name":     l.client_name,
                "client_phone":    l.client_phone,
                "site_status":     l.site_status,
                "stage":           l.stage,
                "material":        l.material,
                "quantity":        l.quantity,
                "sales_exec_name": l.sales_exec_name,
                "is_overdue":      _is_overdue(l),
                "created_at":      l.created_at.isoformat() if l.created_at else None,
                "last_updated":    latest_updates[l.id].isoformat()
                                   if l.id in latest_updates
                                   else (l.created_at.isoformat() if l.created_at else None),
            } for l in recent],
        })
    finally:
        session.close()


@app.get("/admin/api/leads")
def admin_leads(_auth: None = Depends(require_admin)) -> JSONResponse:
    from sqlalchemy import func
    session = SessionLocal()
    try:
        leads    = session.query(Lead).order_by(Lead.created_at.desc()).limit(500).all()
        lead_ids = [l.id for l in leads]
        raw = (
            session.query(
                LeadUpdate.lead_id,
                func.count(LeadUpdate.id).label("cnt"),
                func.max(LeadUpdate.updated_at).label("latest"),
            )
            .filter(LeadUpdate.lead_id.in_(lead_ids))
            .group_by(LeadUpdate.lead_id).all()
        ) if lead_ids else []
        upd_counts = {
            r.lead_id: {"count": r.cnt, "last_updated": r.latest.isoformat() if r.latest else None}
            for r in raw
        }

        return JSONResponse({"leads": [{
            "id":              l.id,
            "company_name":    l.company_name,
            "client_name":     l.client_name,
            "client_phone":    l.client_phone,
            "site_status":     l.site_status,
            "stage":           l.stage,
            "material":        l.material,
            "grade":           l.grade,
            "quantity":        l.quantity,
            "remarks":         l.remarks,
            "location":        l.location,
            "sales_exec_name": l.sales_exec_name,
            "sales_exec_id":   l.sales_exec_id,
            "is_overdue":      _is_overdue(l),
            "created_at":      l.created_at.isoformat() if l.created_at else None,
            "update_count":    upd_counts.get(l.id, {}).get("count", 0),
            "last_updated":    upd_counts.get(l.id, {}).get("last_updated")
                               or (l.created_at.isoformat() if l.created_at else None),
        } for l in leads]})
    finally:
        session.close()


@app.get("/admin/api/lead/{lead_id}")
def admin_lead_detail(
    lead_id: int,
    _auth: None = Depends(require_admin),
) -> JSONResponse:
    session = SessionLocal()
    try:
        lead = session.query(Lead).filter(Lead.id == lead_id).first()
        if not lead:
            raise HTTPException(404, "Lead not found")
        updates = (session.query(LeadUpdate)
                   .filter(LeadUpdate.lead_id == lead_id)
                   .order_by(LeadUpdate.updated_at.desc()).all())
        return JSONResponse({
            "id":              lead.id,
            "company_name":    lead.company_name,
            "client_name":     lead.client_name,
            "client_phone":    lead.client_phone,
            "site_status":     lead.site_status,
            "stage":           lead.stage,
            "material":        lead.material,
            "grade":           lead.grade,
            "quantity":        lead.quantity,
            "remarks":         lead.remarks,
            "location":        lead.location,
            "latitude":        lead.latitude,
            "longitude":       lead.longitude,
            "sales_exec_name": lead.sales_exec_name,
            "sales_exec_id":   lead.sales_exec_id,
            "is_overdue":      _is_overdue(lead),
            "photo_paths":     lead.photo_paths.split(",") if lead.photo_paths else [],
            "created_at":      lead.created_at.isoformat()          if lead.created_at          else None,
            "last_followup_at":lead.last_followup_at.isoformat()    if lead.last_followup_at    else None,
            "next_followup_date":lead.next_followup_date.isoformat() if lead.next_followup_date else None,
            "updates": [{
                "id":          u.id,
                "site_status": u.site_status,
                "stage":       u.stage,
                "material":    u.material,
                "quantity":    u.quantity,
                "remarks":     u.remarks,
                "updated_at":  u.updated_at.isoformat() if u.updated_at else None,
            } for u in updates],
        })
    finally:
        session.close()


@app.post("/admin/api/lead/{lead_id}")
async def admin_update_lead(
    lead_id: int,
    _auth: None = Depends(require_admin),
    request: Request = None,
) -> JSONResponse:
    data = await request.json()
    session = SessionLocal()
    try:
        lead = session.query(Lead).filter(Lead.id == lead_id).first()
        if not lead:
            raise HTTPException(404, "Lead not found")
        for field in ("company_name", "site_status", "stage", "material",
                      "grade", "quantity", "remarks"):
            if field in data:
                setattr(lead, field, data[field])
        if "contact_name" in data: lead.client_name  = data["contact_name"]
        if "phone"        in data: lead.client_phone = data["phone"]
        session.commit()
        return JSONResponse({"status": "ok"})
    except HTTPException:
        raise
    except Exception as exc:
        session.rollback()
        raise HTTPException(500, str(exc))
    finally:
        session.close()


# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN API — REASSIGN / EXEC LIST / ACTIVITY / EXECS / TARGETS / PROGRESS
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/admin/api/lead/{lead_id}/reassign")
async def admin_reassign_lead(
    lead_id: int,
    _auth: None = Depends(require_admin),
    request: Request = None,
) -> JSONResponse:
    data = await request.json()
    new_exec_id   = data.get("sales_exec_id")
    new_exec_name = (data.get("sales_exec_name") or "").strip()
    if not new_exec_id or not new_exec_name:
        raise HTTPException(400, "sales_exec_id and sales_exec_name are required")

    session = SessionLocal()
    try:
        lead = session.query(Lead).filter(Lead.id == lead_id).first()
        if not lead:
            raise HTTPException(404, "Lead not found")

        old_name             = lead.sales_exec_name
        lead.sales_exec_id   = int(new_exec_id)
        lead.sales_exec_name = new_exec_name

        snapshot = LeadUpdate(
            lead_id         = lead.id,
            sales_exec_id   = lead.sales_exec_id,
            sales_exec_name = lead.sales_exec_name,
            company_name    = lead.company_name,
            client_name     = lead.client_name,
            client_phone    = lead.client_phone,
            site_status     = lead.site_status,
            stage           = lead.stage,
            material        = lead.material,
            grade           = lead.grade,
            quantity        = lead.quantity,
            remarks         = f"⚠️ REASSIGNED by Admin (from {old_name} to {new_exec_name})",
        )
        session.add(snapshot)
        session.commit()

        reschedule_on_update(lead.id, lead.sales_exec_id, lead.stage, telegram_app.bot, bot_loop)

        try:
            base_url   = (os.getenv("MINIAPP_BASE_URL") or os.getenv("WEBHOOK_URL") or "").rstrip("/")
            update_url = f"{base_url}/miniapp/update?lead_id={lead.id}&user_id={lead.sales_exec_id}"
            kb = InlineKeyboardMarkup([[
                InlineKeyboardButton("📝 Update Submission", web_app=WebAppInfo(url=update_url))
            ]])
            text = (
                f"🔀 <b>Lead Reassigned to You!</b>\n\n"
                f"Hi {new_exec_name},\n"
                f"Admin has transferred this lead to you:\n\n"
                f"🏢 <b>{lead.company_name}</b>\n"
                f"👤 <b>{lead.client_name}</b> ({lead.client_phone})\n"
                f"📍 <b>Stage:</b> {lead.stage}\n"
                f"⏳ <b>Status:</b> {lead.site_status}"
            )
            async def _send(cid, t, k):
                await telegram_app.bot.send_message(chat_id=cid, text=t, parse_mode="HTML", reply_markup=k)
            asyncio.run_coroutine_threadsafe(_send(lead.sales_exec_id, text, kb), bot_loop)
        except Exception as exc:
            print(f"⚠️ Notification error: {exc}")

        return JSONResponse({"status": "ok", "lead_id": lead.id})
    except HTTPException:
        raise
    except Exception as exc:
        session.rollback()
        raise HTTPException(500, str(exc))
    finally:
        session.close()


@app.get("/admin/api/exec_list")
def admin_exec_list(_auth: None = Depends(require_admin)) -> JSONResponse:
    session = SessionLocal()
    try:
        rows = session.query(Lead.sales_exec_id, Lead.sales_exec_name).distinct().all()
        return JSONResponse({"execs": [{"id": str(r[0]), "name": r[1]} for r in rows if r[0]]})
    finally:
        session.close()


@app.get("/admin/api/activity")
def admin_activity(_auth: None = Depends(require_admin)) -> JSONResponse:
    session = SessionLocal()
    try:
        leads   = session.query(Lead).order_by(Lead.created_at.desc()).limit(200).all()
        updates = session.query(LeadUpdate).order_by(LeadUpdate.updated_at.desc()).limit(200).all()
        lead_map = {l.id: l for l in leads}

        events = []
        for l in leads:
            events.append({
                "type": "lead", "lead_id": l.id,
                "company_name": l.company_name, "client_name": l.client_name,
                "site_status": l.site_status, "stage": l.stage,
                "material": l.material, "quantity": l.quantity, "remarks": l.remarks,
                "sales_exec_name": l.sales_exec_name,
                "created_at": l.created_at.isoformat() if l.created_at else None,
            })
        for u in updates:
            p = lead_map.get(u.lead_id)
            events.append({
                "type": "update", "lead_id": u.lead_id,
                "company_name": p.company_name if p else "—",
                "client_name":  p.client_name  if p else "—",
                "site_status": u.site_status, "stage": u.stage,
                "material": u.material, "quantity": u.quantity, "remarks": u.remarks,
                "sales_exec_name": u.sales_exec_name,
                "created_at": u.updated_at.isoformat() if u.updated_at else None,
            })

        events.sort(key=lambda e: e["created_at"] or "", reverse=True)
        return JSONResponse({"events": events[:300]})
    finally:
        session.close()


@app.get("/admin/api/execs")
def admin_execs(_auth: None = Depends(require_admin)) -> JSONResponse:
    from sqlalchemy import func
    session = SessionLocal()
    try:
        rows = (session.query(
                    Lead.sales_exec_id, Lead.sales_exec_name,
                    func.count(Lead.id).label("total"),
                    func.max(Lead.created_at).label("last_active"),
                )
                .group_by(Lead.sales_exec_id, Lead.sales_exec_name)
                .order_by(func.count(Lead.id).desc()).all())

        result = []
        for r in rows:
            q    = session.query(Lead).filter(Lead.sales_exec_id == r.sales_exec_id)
            won  = q.filter(Lead.site_status == "Won").count()
            lost = q.filter(Lead.site_status == "Lost").count()
            result.append({
                "exec_id":     r.sales_exec_id,
                "name":        r.sales_exec_name or f"Exec #{r.sales_exec_id}",
                "total":       r.total,
                "won":         won,
                "lost":        lost,
                "in_progress": max(r.total - won - lost, 0),
                "last_active": r.last_active.isoformat() if r.last_active else None,
            })
        return JSONResponse({"execs": result})
    finally:
        session.close()


@app.api_route("/admin/api/execs/{exec_id}/target", methods=["GET", "POST"])
async def admin_exec_target(
    exec_id: int,
    request: Request,
    _auth: None = Depends(require_admin),
) -> JSONResponse:
    session = SessionLocal()
    try:
        target    = session.query(ExecTarget).filter(ExecTarget.sales_exec_id == exec_id).first()
        exec_name = (
            session.query(Lead.sales_exec_name)
            .filter(Lead.sales_exec_id == exec_id).limit(1).scalar()
            or f"Exec #{exec_id}"
        )

        if request.method == "GET":
            return JSONResponse({
                "exec_id":        exec_id,
                "exec_name":      target.sales_exec_name if target else exec_name,
                "monthly_leads":  target.monthly_leads   if target else 30,
                "conversion_pct": target.conversion_pct  if target else 40.0,
                "volume_m3":      target.volume_m3        if target else 500.0,
            })

        data = await request.json()
        if not target:
            target = ExecTarget(sales_exec_id=exec_id, sales_exec_name=exec_name)
            session.add(target)
        target.sales_exec_name = exec_name
        if "monthly_leads"  in data: target.monthly_leads  = int(data["monthly_leads"])
        if "conversion_pct" in data: target.conversion_pct = float(data["conversion_pct"])
        if "volume_m3"      in data: target.volume_m3      = float(data["volume_m3"])
        session.commit()
        return JSONResponse({"status": "ok"})
    except HTTPException:
        raise
    except Exception as exc:
        session.rollback()
        raise HTTPException(500, str(exc))
    finally:
        session.close()


@app.get("/admin/api/execs/{exec_id}/progress")
def admin_exec_progress(
    exec_id: int,
    _auth: None = Depends(require_admin),
) -> JSONResponse:
    from sqlalchemy import func as _func
    session = SessionLocal()
    try:
        now_ist     = datetime.now(timezone.utc) + IST_OFFSET
        month_start = now_ist.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        month_start_utc = month_start - IST_OFFSET
        week_start  = now_ist - timedelta(days=now_ist.weekday())
        week_start  = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start_utc = week_start - IST_OFFSET

        q_all   = session.query(Lead).filter(Lead.sales_exec_id == exec_id)
        q_month = q_all.filter(Lead.created_at >= month_start_utc)
        q_week  = q_all.filter(Lead.created_at >= week_start_utc)

        def _breakdown(q):
            won = q.filter(Lead.site_status == "Won").count()
            lost= q.filter(Lead.site_status == "Lost").count()
            neg = q.filter(Lead.site_status == "Negotiation in Progress").count()
            vis = q.filter(Lead.site_status == "Visited").count()
            quo = q.filter(Lead.site_status == "Quoted").count()
            total = won + lost + neg + vis + quo
            volume = sum(
                float(l.quantity) for l in q.filter(Lead.quantity.isnot(None)).all()
                if _safe_float(l.quantity) is not None
            )
            return {
                "won": won, "lost": lost, "negotiating": neg,
                "visited": vis, "quoted": quo, "total": total,
                "volume": round(volume, 2),
                "conv_pct": round(won / total * 100, 1) if total else 0,
            }

        monthly = _breakdown(q_month)
        weekly  = _breakdown(q_week)

        stage_rows = (
            session.query(Lead.stage, _func.count(Lead.id))
            .filter(Lead.sales_exec_id == exec_id, Lead.created_at >= month_start_utc)
            .group_by(Lead.stage).all()
        )
        stage_breakdown = [{"stage": s or "—", "count": c} for s, c in stage_rows]

        thirty_ago     = now_ist - timedelta(days=29)
        thirty_ago_utc = thirty_ago.replace(hour=0, minute=0, second=0, microsecond=0) - IST_OFFSET
        daily_rows = (
            session.query(Lead.created_at)
            .filter(Lead.sales_exec_id == exec_id, Lead.created_at >= thirty_ago_utc)
            .all()
        )
        daily_map: dict = {}
        for (dt,) in daily_rows:
            d = (dt + IST_OFFSET).strftime("%Y-%m-%d")
            daily_map[d] = daily_map.get(d, 0) + 1
        daily = [
            {"date": d, "count": daily_map.get(d, 0)}
            for d in [(thirty_ago + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(30)]
        ]

        target    = session.query(ExecTarget).filter(ExecTarget.sales_exec_id == exec_id).first()
        exec_name = (
            session.query(Lead.sales_exec_name).filter(Lead.sales_exec_id == exec_id).limit(1).scalar()
            or f"Exec #{exec_id}"
        )
        return JSONResponse({
            "exec_id":        exec_id,
            "exec_name":      exec_name,
            "monthly":        monthly,
            "weekly":         weekly,
            "stage_breakdown":stage_breakdown,
            "daily_trend":    daily,
            "targets": {
                "monthly_leads":  target.monthly_leads  if target else 30,
                "conversion_pct": target.conversion_pct if target else 40.0,
                "volume_m3":      target.volume_m3      if target else 500.0,
            },
        })
    finally:
        session.close()


# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN API — EXPORTS (CSV / EXCEL)
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/admin/api/export/csv")
def admin_export_csv(_auth: None = Depends(require_admin)) -> StreamingResponse:
    from sqlalchemy import func
    session = SessionLocal()
    try:
        leads    = session.query(Lead).order_by(Lead.created_at.asc()).all()
        lead_ids = [l.id for l in leads]
        latest = dict(
            session.query(LeadUpdate.lead_id, func.max(LeadUpdate.updated_at))
            .filter(LeadUpdate.lead_id.in_(lead_ids))
            .group_by(LeadUpdate.lead_id).all()
        ) if lead_ids else {}
        counts = dict(
            session.query(LeadUpdate.lead_id, func.count(LeadUpdate.id))
            .filter(LeadUpdate.lead_id.in_(lead_ids))
            .group_by(LeadUpdate.lead_id).all()
        ) if lead_ids else {}

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "ID", "Company", "Contact Person", "Phone",
            "Work Status", "Stage", "Material", "Grade", "Quantity",
            "Remarks", "Location", "Sales Executive",
            "First Submitted (IST)", "Last Updated (IST)", "Total Updates",
        ])
        for l in leads:
            last_upd = latest.get(l.id)
            writer.writerow([
                l.id, l.company_name, l.client_name, l.client_phone,
                l.site_status, l.stage, l.material, l.grade or "", l.quantity,
                l.remarks or "", l.location or "", l.sales_exec_name,
                _to_ist_str(l.created_at),
                _to_ist_str(last_upd) if last_upd else _to_ist_str(l.created_at),
                counts.get(l.id, 0),
            ])
        output.seek(0)
        date_str = (datetime.now(timezone.utc) + IST_OFFSET).strftime("%Y-%m-%d")
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=leads_{date_str}.csv"},
        )
    finally:
        session.close()


@app.get("/admin/api/export/excel")
def admin_export_excel(_auth: None = Depends(require_admin)) -> StreamingResponse:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    from sqlalchemy import func
    session = SessionLocal()
    try:
        leads    = session.query(Lead).order_by(Lead.created_at.asc()).all()
        lead_ids = [l.id for l in leads]
        latest = dict(
            session.query(LeadUpdate.lead_id, func.max(LeadUpdate.updated_at))
            .filter(LeadUpdate.lead_id.in_(lead_ids))
            .group_by(LeadUpdate.lead_id).all()
        ) if lead_ids else {}
        counts = dict(
            session.query(LeadUpdate.lead_id, func.count(LeadUpdate.id))
            .filter(LeadUpdate.lead_id.in_(lead_ids))
            .group_by(LeadUpdate.lead_id).all()
        ) if lead_ids else {}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads"

        headers    = ["ID","Company","Contact Person","Phone","Work Status","Stage",
                      "Material","Grade","Quantity","Remarks","Location","Sales Executive",
                      "First Submitted (IST)","Last Updated (IST)","Total Updates"]
        col_widths = [6,24,20,14,20,12,16,10,12,28,22,20,22,22,14]

        header_fill  = PatternFill("solid", fgColor="1A237E")
        header_font  = Font(bold=True, color="FFFFFF", size=11)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border  = Border(
            left  =Side(style="thin", color="E8EAEF"),
            right =Side(style="thin", color="E8EAEF"),
            bottom=Side(style="thin", color="E8EAEF"),
        )
        status_fills = {
            "Won":                     PatternFill("solid", fgColor="E8F5E9"),
            "Lost":                    PatternFill("solid", fgColor="FFEBEE"),
            "Negotiation in Progress": PatternFill("solid", fgColor="FFF3E0"),
            "Visited":                 PatternFill("solid", fgColor="E8EAF6"),
            "Quoted":                  PatternFill("solid", fgColor="E8EAF6"),
        }

        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = header_fill; cell.font = header_font
            cell.alignment = center_align; cell.border = thin_border
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 36
        ws.freeze_panes = "A2"

        alt_fill = PatternFill("solid", fgColor="F7F8FC")
        for ri, l in enumerate(leads, 2):
            last_upd = latest.get(l.id)
            row_data = [
                l.id, l.company_name, l.client_name, l.client_phone,
                l.site_status, l.stage, l.material, l.grade or "", l.quantity or "",
                l.remarks or "", l.location or "", l.sales_exec_name or "",
                _to_ist_str(l.created_at),
                _to_ist_str(last_upd) if last_upd else _to_ist_str(l.created_at),
                counts.get(l.id, 0),
            ]
            row_fill = status_fills.get(l.site_status, alt_fill if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF"))
            for ci, v in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.fill = row_fill; cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[ri].height = 20

        ws2 = wb.create_sheet("Summary")
        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 14
        total = len(leads)
        won   = sum(1 for l in leads if l.site_status == "Won")
        lost  = sum(1 for l in leads if l.site_status == "Lost")
        for ri, (label, value) in enumerate([
            ("OLM — Lead Export", ""),
            ("Generated (IST)", _to_ist_str(datetime.now(timezone.utc))),
            ("", ""),
            ("Total Leads", total), ("Won", won), ("Lost", lost),
            ("In Progress", total - won - lost),
            ("Conversion Rate", f"{round(won/total*100,1)}%" if total else "0%"),
        ], 1):
            ws2.cell(row=ri, column=1, value=label).font = Font(bold=(ri in (1,4,5,6,7,8)))
            ws2.cell(row=ri, column=2, value=value)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        date_str = (datetime.now(timezone.utc) + IST_OFFSET).strftime("%Y-%m-%d")
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=leads_{date_str}.xlsx"},
        )
    finally:
        session.close()


# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN API — UTILITIES (digest / scheduler / health)
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/admin/api/test-digest")
def admin_test_digest(_auth: None = Depends(require_admin)) -> JSONResponse:
    from bot.daily_digest import send_daily_digests
    try:
        send_daily_digests(bot=telegram_app.bot, loop=bot_loop)
        owner = os.getenv("DIGEST_OWNER_CHAT_ID", "").strip()
        return JSONResponse({
            "status": "ok",
            "recipients": "all execs + owner" if owner else "all execs only",
        })
    except Exception as exc:
        raise HTTPException(500, str(exc))


@app.get("/admin/api/scheduler-status")
def admin_scheduler_status(_auth: None = Depends(require_admin)) -> JSONResponse:
    jobs = [
        {"id": j.id, "next_run": j.next_run_time.isoformat() if j.next_run_time else None}
        for j in apscheduler.get_jobs()
    ]
    return JSONResponse({"job_count": len(jobs), "jobs": jobs})


@app.get("/health")
def health_check() -> JSONResponse:
    return JSONResponse({"status": "ok", "version": "2.0.0"})
