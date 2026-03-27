# app.py
from flask import Flask, request, jsonify, send_from_directory
from telegram import Update
from telegram.ext import (Application, CommandHandler, CallbackQueryHandler)
from bot.handlers.lead import start_lead
from bot.handlers.followup import handle_followup_response
from bot.handlers.admin import register_admin_handlers
from models.lead import Lead
from models.lead_update import LeadUpdate
from models.exec_target import ExecTarget
from bot.scheduler import (
    schedule_followups, reschedule_on_update, sync_scheduler_with_db,
    scheduler as apscheduler, print_scheduler_status
)
from bot.daily_digest import register_daily_digest
from db import engine, Base, SessionLocal
from config import TELEGRAM_TOKEN, WEBHOOK_URL
import asyncio, threading
from datetime import datetime, timedelta, timezone

app = Flask(__name__)
Base.metadata.create_all(engine)

telegram_app = Application.builder().token(TELEGRAM_TOKEN).build()

# ── Handlers ──────────────────────────────────────────────────────────────────
# /newlead now sends a Mini App button instead of chat prompts
telegram_app.add_handler(CommandHandler("newlead", start_lead))
telegram_app.add_handler(CommandHandler("start", start_lead))
telegram_app.add_handler(CallbackQueryHandler(handle_followup_response, pattern='^fu_'))
register_admin_handlers(telegram_app)

# ── Background event loop ─────────────────────────────────────────────────────
# Bot's async work runs on this loop; Flask stays synchronous on the main thread
def start_loop(loop):
    asyncio.set_event_loop(loop)
    loop.run_forever()

loop = asyncio.new_event_loop()
threading.Thread(target=start_loop, args=(loop,), daemon=True).start()

async def init_bot():
    await telegram_app.initialize()
    await telegram_app.start()
    await telegram_app.bot.set_webhook(f"{WEBHOOK_URL}/webhook")
    print("Telegram bot initialized & webhook set")

# Block until bot is ready before Flask starts accepting traffic
asyncio.run_coroutine_threadsafe(init_bot(), loop).result(timeout=15)

# Register daily digest cron (runs every day at DIGEST_TIME_IST, default 20:00 IST)
register_daily_digest(apscheduler, telegram_app.bot, loop)

# Sync scheduler with DB at startup to recover any jobs for active leads
sync_scheduler_with_db(telegram_app.bot, loop)

# Periodically sync scheduler with DB (every hour) to ensure no lead is missed
apscheduler.add_job(
    sync_scheduler_with_db,
    trigger="interval",
    hours=1,
    args=[telegram_app.bot, loop],
    id="sync_scheduler",
    replace_existing=True
)

# Print all scheduled follow-up jobs at startup
print_scheduler_status()

# ── Webhook (Telegram → Flask) ────────────────────────────────────────────────
@app.route("/webhook", methods=["POST"])
def webhook():
    update = Update.de_json(request.get_json(force=True), telegram_app.bot)
    asyncio.run_coroutine_threadsafe(
        telegram_app.process_update(update),
        loop
    )
    return "OK", 200

# ── Serve Mini App HTML ───────────────────────────────────────────────────────
@app.route("/miniapp")
def serve_miniapp():
    """Inject MINIAPP_BASE_URL from env into the HTML before serving."""
    import os
    base_url = os.getenv("MINIAPP_BASE_URL", "").rstrip("/")
    with open(os.path.join("miniapp", "index.html"), "r") as f:
        html = f.read()
    # Replace the placeholder the frontend uses for the API base URL
    html = html.replace("https://YOUR_NGROK_OR_DOMAIN", base_url)
    return html, 200, {"Content-Type": "text/html; charset=utf-8", "Cache-Control": "no-store, no-cache, must-revalidate"}

# ── Receive lead from Mini App form submission ────────────────────────────────
@app.route("/submit_lead", methods=["POST"])
def submit_lead():
    import json, os
    from werkzeug.utils import secure_filename

    # Mini App sends multipart/form-data (not JSON) because it includes photos
    f = request.form

    company_name = f.get("company_name", "").strip()
    contact_name = f.get("contact_name", "").strip()
    phone        = f.get("phone", "").strip()
    work_status  = f.get("work_status", "").strip()
    stage        = f.get("stage", "").strip()
    material     = f.get("material", "").strip()
    grade        = f.get("grade", "").strip()
    quantity     = f.get("quantity", "").strip()
    remarks      = f.get("remarks", "").strip()

    try:
        latitude  = float(f.get("latitude"))  if f.get("latitude")  else None
        longitude = float(f.get("longitude")) if f.get("longitude") else None
    except (ValueError, TypeError):
        latitude = longitude = None

    try:
        tg_user = json.loads(f.get("tg_user", "{}"))
    except (ValueError, TypeError):
        tg_user = {}
    chat_id   = tg_user.get("id")
    user_name = tg_user.get("name") or \
                " ".join(filter(None, [tg_user.get("first_name",""), tg_user.get("last_name","")])) or \
                tg_user.get("username") or "Unknown"

    # Validate required fields
    missing = [k for k, v in {
        "company_name": company_name,
        "contact_name": contact_name,
        "phone":        phone,
        "work_status":  work_status,
        "stage":        stage,
        "material":     material,
        "quantity":     quantity,
    }.items() if not v]

    if missing:
        return jsonify({"error": f"Missing required fields: {', '.join(missing)}"}), 400

    if not latitude or not longitude:
        return jsonify({"error": "Location is required"}), 400

    # Save uploaded photos to disk
    UPLOAD_DIR = os.path.join("uploads", "leads")
    os.makedirs(UPLOAD_DIR, exist_ok=True)

    photo_paths = []
    for key in request.files:
        if key.startswith("photo_"):
            file = request.files[key]
            if file and file.filename:
                filename = secure_filename(f"{chat_id or 'unknown'}_{key}_{file.filename}")
                save_path = os.path.join(UPLOAD_DIR, filename)
                file.save(save_path)
                photo_paths.append(save_path)

    if not photo_paths:
        return jsonify({"error": "At least one photo is required"}), 400

    session = SessionLocal()
    try:
        lead = Lead(
            company_name=company_name,
            client_name=contact_name,
            client_phone=phone,
            site_status=work_status,
            stage=stage,
            material=material,
            grade=grade,
            quantity=quantity,
            remarks=remarks,
            photo_paths=",".join(photo_paths),
            latitude=latitude,
            longitude=longitude,
            location=f"{latitude:.5f},{longitude:.5f}",
            sales_exec_id=chat_id,
            sales_exec_name=user_name,        # ← saved for display in admin/history
        )
        session.add(lead)
        session.commit()
        session.refresh(lead)

        if chat_id:
            schedule_followups(
                lead_id=lead.id,
                chat_id=chat_id,
                bot=telegram_app.bot,
                loop=loop,
                stage=stage,       # drives Pile=daily, Footing/Slab=3d, Flooring=4d, Column=7d
            )

        print(f"Lead #{lead.id} saved — {company_name} | {stage} | {material}")
        return jsonify({"status": "ok", "lead_id": lead.id}), 200

    except Exception as e:
        session.rollback()
        print(f"DB error: {e}")
        return jsonify({"error": "Database error"}), 500
    finally:
        session.close()

# ── Serve Mini App Update Form ────────────────────────────────────────────────
@app.route("/miniapp/update")
def serve_miniapp_update():
    """Same index.html — the JS reads ?lead_id= and switches to update mode."""
    import os
    base_url = os.getenv("MINIAPP_BASE_URL", "").rstrip("/")
    with open(os.path.join("miniapp", "index.html"), "r") as f:
        html = f.read()
    html = html.replace("https://YOUR_NGROK_OR_DOMAIN", base_url)
    return html, 200, {"Content-Type": "text/html; charset=utf-8", "Cache-Control": "no-store, no-cache, must-revalidate"}

# ── Generate Quote PDF and send to Telegram ───────────────────────────────────
@app.route("/generate_quote", methods=["POST"])
def generate_quote():
    """
    Accepts JSON: company_name, quantity, grade, rate, location, tg_user.
    Generates a 4-page Mcube quote PDF (page 1 dynamic, pages 2-4 from template)
    and sends it to the user's Telegram chat.
    """
    import io, os, json as _json

    data     = request.get_json(force=True) or {}
    company  = data.get("company_name", "").strip()
    quantity = data.get("quantity",     "").strip()
    grade    = data.get("grade",        "").strip()
    rate     = data.get("rate",         "").strip()
    location = data.get("location",     "").strip()
    user_id  = data.get("user_id")

    try:
        tg_user = _json.loads(data.get("tg_user", "{}")) \
                  if isinstance(data.get("tg_user"), str) \
                  else data.get("tg_user", {})
    except Exception:
        tg_user = {}

    chat_id   = tg_user.get("id") or user_id
    exec_name = (
        " ".join(filter(None, [
            tg_user.get("first_name", ""),
            tg_user.get("last_name",  "")
        ]))
        or tg_user.get("username")
        or "Executive"
    ).strip()

    if not all([company, quantity, grade, rate, location, chat_id]):
        return jsonify({"error": "Missing required fields"}), 400

    # Path to the original template PDF (pages 2-4 reused verbatim)
    TEMPLATE_PATH = os.path.join(
        os.getenv("QUOTE_TEMPLATE_DIR", "static"),
        "quote_template.pdf"
    )
    if not os.path.exists(TEMPLATE_PATH):
        return jsonify({
            "error": (
                f"Quote template not found at '{TEMPLATE_PATH}'. "
                "Copy the original quote PDF to static/quote_template.pdf"
            )
        }), 500

    from bot.quote_generator import build_quote_pdf
    buf = io.BytesIO()
    try:
        build_quote_pdf(
            out_buf       = buf,
            template_path = TEMPLATE_PATH,
            company       = company,
            location      = location,
            quantity      = quantity,
            grade         = grade,
            rate          = rate,
            exec_name     = exec_name,
        )
        buf.seek(0)
    except Exception as e:
        print(f"Quote build error: {e}")
        return jsonify({"error": f"PDF generation failed: {e}"}), 500

    now_ist  = datetime.now(timezone(timedelta(hours=5, minutes=30)))
    date_str = now_ist.strftime("%d %b %Y")
    filename = f"Mcube_Quote_{company.replace(' ','_')}_{date_str.replace(' ','')}.pdf"

    try:
        qty_f  = float(quantity)
    except (ValueError, TypeError):
        qty_f  = 0.0
    try:
        rate_f = float(rate)
    except (ValueError, TypeError):
        rate_f = 0.0

    async def _send_pdf():
        await telegram_app.bot.send_document(
            chat_id  = int(chat_id),
            document = buf,
            filename = filename,
            caption  = (
                f"📄 *Quote — {company}*\n"
                f"📍 {location}  |  🏗 {grade}  |  📦 {qty_f:.0f} cum\n"
                f"💰 Rate: \u20b9{rate_f:,.0f}/cum (incl. GST)"
            ),
            parse_mode = "Markdown",
        )

    future = asyncio.run_coroutine_threadsafe(_send_pdf(), loop)
    try:
        # Increased timeout to 30s to handle network latency or larger PDFs
        future.result(timeout=30)
        print(f"Quote sent to {chat_id} — {company} | {grade} | {qty_f:.0f} cum")
        return jsonify({"status": "ok", "filename": filename})
    except Exception as e:
        print(f"Quote send error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/auth", methods=["POST"])
def api_auth():
    """Validate miniapp access key. Key is set via MINIAPP_ACCESS_KEY in .env."""
    import os as _os
    data = request.get_json(force=True) or {}
    key  = (data.get("key") or "").strip()

    access_key = _os.getenv("MINIAPP_ACCESS_KEY", "").strip()
    if not access_key:
        # No key configured — allow all (open access)
        return jsonify({"status": "ok"})

    if key == access_key:
        return jsonify({"status": "ok"})

    return jsonify({"error": "Invalid access key"}), 401


@app.route("/api/lead/<int:lead_id>")
def api_lead_detail(lead_id):
    user_id = request.args.get("user_id", type=int)
    if not user_id:
        return jsonify({"error": "user_id is required"}), 400

    session = SessionLocal()
    try:
        lead = (session.query(Lead)
                .filter(Lead.id == lead_id, Lead.sales_exec_id == user_id)
                .first())
        if not lead:
            return jsonify({"error": "Lead not found"}), 404

        return jsonify({
            "id":           lead.id,
            "company_name": lead.company_name,
            "client_name":  lead.client_name,
            "client_phone": lead.client_phone,
            "site_status":  lead.site_status,
            "stage":        lead.stage,
            "material":     lead.material,
            "grade":        lead.grade,
            "quantity":     lead.quantity,
            "remarks":      lead.remarks,
            "latitude":     lead.latitude,
            "longitude":    lead.longitude,
        })
    finally:
        session.close()

# ── API: Update lead ──────────────────────────────────────────────────────────
@app.route("/api/lead/<int:lead_id>/update", methods=["POST"])
def api_update_lead(lead_id):
    import json as _json
    f = request.form
    user_id   = None
    user_name = "Unknown"
    try:
        tg_user   = _json.loads(f.get("tg_user", "{}"))
        user_id   = tg_user.get("id")
        user_name = tg_user.get("name") or \
                    " ".join(filter(None, [tg_user.get("first_name",""), tg_user.get("last_name","")])) or \
                    tg_user.get("username") or "Unknown"
    except Exception:
        pass

    if not user_id:
        return jsonify({"error": "user_id is required"}), 400

    session = SessionLocal()
    try:
        lead = (session.query(Lead)
                .filter(Lead.id == lead_id, Lead.sales_exec_id == user_id)
                .first())
        if not lead:
            return jsonify({"error": "Lead not found"}), 404

        # Apply updates to lead
        if f.get("company_name"):  lead.company_name = f.get("company_name").strip()
        if f.get("contact_name"):  lead.client_name  = f.get("contact_name").strip()
        if f.get("phone"):         lead.client_phone = f.get("phone").strip()
        if f.get("work_status"):   lead.site_status  = f.get("work_status").strip()
        if f.get("stage"):         lead.stage        = f.get("stage").strip()
        if f.get("material"):      lead.material     = f.get("material").strip()
        if f.get("grade"):         lead.grade        = f.get("grade").strip()
        if f.get("quantity"):      lead.quantity     = f.get("quantity").strip()
        if f.get("remarks"):       lead.remarks      = f.get("remarks").strip()

        # Save update snapshot for history timeline
        new_stage = lead.stage   # capture after potential update above
        snapshot = LeadUpdate(
            lead_id         = lead_id,
            sales_exec_id   = user_id,
            sales_exec_name = user_name,
            company_name    = lead.company_name,
            client_name     = lead.client_name,
            client_phone    = lead.client_phone,
            site_status     = lead.site_status,
            stage           = new_stage,
            material        = lead.material,
            grade           = lead.grade,
            quantity        = lead.quantity,
            remarks         = lead.remarks,
        )
        session.add(snapshot)
        session.commit()
        
        # Reset follow-up timer on any update — logic now checks DB times
        reschedule_on_update(lead.id, user_id, lead.stage, telegram_app.bot, loop)
        
        print(f"Lead #{lead_id} updated by {user_name} ({user_id})")
        return jsonify({"status": "ok", "lead_id": lead.id}), 200

    except Exception as e:
        session.rollback()
        print(f"Update error: {e}")
        return jsonify({"error": "Database error"}), 500
    finally:
        session.close()


# ── TEST: Trigger reminder instantly — remove in production ───────────────────
@app.route("/testmessage")
def test_reminder():
    from bot.scheduler import send_followup_reminder
    user_id = request.args.get("user_id", type=int)
    lead_id = request.args.get("lead_id", type=int)

    if not user_id:
        return jsonify({"error": "user_id is required"}), 400
    if not lead_id:
        return jsonify({"error": "lead_id is required"}), 400

    session = SessionLocal()
    try:
        lead = session.query(Lead).filter(
            Lead.id == lead_id,
            Lead.sales_exec_id == user_id
        ).first()
        if not lead:
            return jsonify({"error": f"Lead #{lead_id} not found for user {user_id}"}), 404
    finally:
        session.close()

    try:
        send_followup_reminder(
            lead_id=lead_id,
            chat_id=user_id,
            bot=telegram_app.bot,
            loop=loop,
        )
        return jsonify({
            "status": "Test reminder sent",
            "lead_id": lead_id,
            "chat_id": user_id
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/stats")
def api_stats():
    user_id = request.args.get("user_id", type=int)
    if not user_id:
        return jsonify({"error": "user_id is required"}), 400

    session = SessionLocal()
    try:
        q = session.query(Lead).filter(Lead.sales_exec_id == user_id)

        total       = q.count()
        won         = q.filter(Lead.site_status == "Won").count()
        lost        = q.filter(Lead.site_status == "Lost").count()
        in_progress = max(total - won - lost, 0)

        recent = q.order_by(Lead.created_at.desc()).limit(10).all()

        return jsonify({
            "total":       total,
            "won":         won,
            "lost":        lost,
            "in_progress": in_progress,
            "recent": [{
                "id":           l.id,
                "company_name": l.company_name,
                "client_name":  l.client_name,
                "site_status":  l.site_status,
                "stage":        l.stage,
                "material":     l.material,
                "quantity":     l.quantity,
                "created_at":   l.created_at.isoformat() if l.created_at else None,
            } for l in recent]
        })
    finally:
        session.close()

# ── API: Goals ────────────────────────────────────────────────────────────────
@app.route("/api/goals")
def api_goals():
    from datetime import datetime
    user_id = request.args.get("user_id", type=int)
    if not user_id:
        return jsonify({"error": "user_id is required"}), 400

    session = SessionLocal()
    try:
        q  = session.query(Lead).filter(Lead.sales_exec_id == user_id)
        now   = datetime.now(timezone.utc)
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        qm    = q.filter(Lead.created_at >= start)

        this_month = qm.count()
        won        = qm.filter(Lead.site_status == "Won").count()

        total_qty = 0.0
        for l in qm.filter(Lead.quantity.isnot(None)).all():
            try:    total_qty += float(l.quantity)
            except: pass

        # Read targets from ExecTarget if set, otherwise fall back to defaults
        target = session.query(ExecTarget).filter(
            ExecTarget.sales_exec_id == user_id
        ).first()

        goals = {
            "monthly_target":    target.monthly_leads  if target else 30,
            "conversion_target": target.conversion_pct if target else 40,
            "volume_target":     target.volume_m3      if target else 500,
        }

        return jsonify({
            "this_month":     this_month,
            "won":            won,
            "total_quantity": round(total_qty, 2),
            "goals":          goals,
        })
    finally:
        session.close()

# ── API: History ──────────────────────────────────────────────────────────────
@app.route("/api/history")
def api_history():
    user_id = request.args.get("user_id", type=int)
    if not user_id:
        return jsonify({"error": "user_id is required"}), 400

    session = SessionLocal()
    try:
        leads = (session.query(Lead)
                 .filter(Lead.sales_exec_id == user_id)
                 .order_by(Lead.created_at.desc())
                 .limit(50).all())

        # Fetch all updates for these leads in one query
        lead_ids = [l.id for l in leads]
        updates  = (session.query(LeadUpdate)
                    .filter(LeadUpdate.lead_id.in_(lead_ids))
                    .order_by(LeadUpdate.updated_at.asc())
                    .all()) if lead_ids else []

        # Group updates by lead_id
        updates_by_lead = {}
        for u in updates:
            updates_by_lead.setdefault(u.lead_id, []).append({
                "id":          u.id,
                "site_status": u.site_status,
                "stage":       u.stage,
                "material":    u.material,
                "quantity":    u.quantity,
                "remarks":     u.remarks,
                "updated_at":  u.updated_at.isoformat() if u.updated_at else None,
            })

        return jsonify({
            "leads": [{
                "id":           l.id,
                "company_name": l.company_name,
                "client_name":  l.client_name,
                "client_phone": l.client_phone,
                "site_status":  l.site_status,
                "stage":        l.stage,
                "material":     l.material,
                "grade":        l.grade,
                "quantity":     l.quantity,
                "remarks":      l.remarks,
                "location":     l.location,
                "created_at":   l.created_at.isoformat() if l.created_at else None,
                "updates":      updates_by_lead.get(l.id, []),  # timeline of updates
            } for l in leads]
        })
    finally:
        session.close()


# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN DASHBOARD — /admin routes
# ══════════════════════════════════════════════════════════════════════════════
import os as _os
from functools import wraps

ADMIN_PASSWORD = _os.getenv("ADMIN_PASSWORD", "mcube@admin123")

def require_admin(f):
    """Protect admin API endpoints with X-Admin-Key header OR ?key= query param."""
    @wraps(f)
    def decorated(*args, **kwargs):
        key = (request.headers.get("X-Admin-Key") or
               request.args.get("key") or "")
        if key != ADMIN_PASSWORD:
            return jsonify({"error": "Unauthorized"}), 401
        return f(*args, **kwargs)
    return decorated


@app.route("/")
def home():
    """Serve the admin dashboard HTML with injected config."""
    base_url = _os.getenv("MINIAPP_BASE_URL") or _os.getenv("WEBHOOK_URL", "")
    base_url = base_url.rstrip("/")
    with open(_os.path.join("admin", "dashboard.html"), "r") as f:
        html = f.read()
    html = html.replace("ADMIN_BASE_URL_PLACEHOLDER", base_url)
    html = html.replace("ADMIN_PASSWORD_PLACEHOLDER", ADMIN_PASSWORD)
    return html, 200, {"Content-Type": "text/html; charset=utf-8", "Cache-Control": "no-store, no-cache, must-revalidate"}

@app.route("/admin")
def admin_dashboard():
    """Serve the admin dashboard HTML with injected config."""
    base_url = _os.getenv("MINIAPP_BASE_URL") or _os.getenv("WEBHOOK_URL", "")
    base_url = base_url.rstrip("/")
    with open(_os.path.join("admin", "dashboard.html"), "r") as f:
        html = f.read()
    html = html.replace("ADMIN_BASE_URL_PLACEHOLDER", base_url)
    html = html.replace("ADMIN_PASSWORD_PLACEHOLDER", ADMIN_PASSWORD)
    return html, 200, {"Content-Type": "text/html; charset=utf-8", "Cache-Control": "no-store, no-cache, must-revalidate"}

@app.route("/admin/api/overview")
@require_admin
def admin_overview():
    from sqlalchemy import func
    session = SessionLocal()
    try:
        total       = session.query(Lead).count()
        won         = session.query(Lead).filter(Lead.site_status == "Won").count()
        lost        = session.query(Lead).filter(Lead.site_status == "Lost").count()
        in_progress = max(total - won - lost, 0)
        recent      = session.query(Lead).order_by(Lead.created_at.desc()).limit(15).all()

        # Fetch latest update timestamp per lead in one query
        lead_ids = [l.id for l in recent]
        latest_updates = dict(
            session.query(LeadUpdate.lead_id, func.max(LeadUpdate.updated_at))
            .filter(LeadUpdate.lead_id.in_(lead_ids))
            .group_by(LeadUpdate.lead_id).all()
        ) if lead_ids else {}

        return jsonify({
            "total": total, "won": won, "lost": lost, "in_progress": in_progress,
            "recent": [{
                "id":              l.id,
                "company_name":    l.company_name,
                "client_name":     l.client_name,
                "client_phone":    l.client_phone,
                "site_status":     l.site_status,
                "stage":           l.stage,
                "material":        l.material,
                "quantity":        l.quantity,
                "sales_exec_name": l.sales_exec_name,
                "created_at":      l.created_at.isoformat() if l.created_at else None,
                "last_updated":    latest_updates[l.id].isoformat()
                                   if l.id in latest_updates
                                   else (l.created_at.isoformat() if l.created_at else None),
            } for l in recent]
        })
    finally:
        session.close()

@app.route("/admin/api/leads")
@require_admin
def admin_leads():
    session = SessionLocal()
    try:
        leads = session.query(Lead).order_by(Lead.created_at.desc()).limit(500).all()
        lead_ids = [l.id for l in leads]
        # Count updates per lead efficiently
        from sqlalchemy import func
        raw = (
            session.query(
                LeadUpdate.lead_id,
                func.count(LeadUpdate.id).label("cnt"),
                func.max(LeadUpdate.updated_at).label("latest")
            )
            .filter(LeadUpdate.lead_id.in_(lead_ids))
            .group_by(LeadUpdate.lead_id).all()
        ) if lead_ids else []
        upd_counts = {
            r.lead_id: {
                "count":        r.cnt,
                "last_updated": r.latest.isoformat() if r.latest else None
            } for r in raw
        }
        return jsonify({"leads": [{
            "id": l.id, "company_name": l.company_name, "client_name": l.client_name,
            "client_phone": l.client_phone, "site_status": l.site_status,
            "stage": l.stage, "material": l.material, "quantity": l.quantity,
            "sales_exec_name": l.sales_exec_name,
            "created_at":   l.created_at.isoformat()   if l.created_at   else None,
            "last_updated": upd_counts.get(l.id, {}).get("last_updated") or
                            (l.created_at.isoformat() if l.created_at else None),
            "update_count": upd_counts.get(l.id, {}).get("count", 0),
        } for l in leads]})
    finally:
        session.close()

@app.route("/admin/api/lead/<int:lead_id>")
@require_admin
def admin_lead_detail(lead_id):
    session = SessionLocal()
    try:
        l = session.query(Lead).filter(Lead.id == lead_id).first()
        if not l:
            return jsonify({"error": "Not found"}), 404
        updates = (session.query(LeadUpdate)
                   .filter(LeadUpdate.lead_id == lead_id)
                   .order_by(LeadUpdate.updated_at.asc()).all())
        return jsonify({
            "id": l.id, "company_name": l.company_name, "client_name": l.client_name,
            "client_phone": l.client_phone, "site_status": l.site_status,
            "stage": l.stage, "material": l.material, "grade": l.grade,
            "quantity": l.quantity, "remarks": l.remarks,
            "latitude": l.latitude, "longitude": l.longitude, "location": l.location,
            "sales_exec_name": l.sales_exec_name,
            "created_at":   l.created_at.isoformat() if l.created_at else None,
            "last_updated": updates[-1].updated_at.isoformat() if updates else
                            (l.created_at.isoformat() if l.created_at else None),
            "updates": [{
                "id": u.id, "site_status": u.site_status, "stage": u.stage,
                "material": u.material, "quantity": u.quantity, "remarks": u.remarks,
                "updated_at": u.updated_at.isoformat() if u.updated_at else None,
            } for u in updates]
        })
    finally:
        session.close()

@app.route("/admin/api/activity")
@require_admin
def admin_activity():
    """Merge lead submissions and updates into a single timeline, newest first."""
    session = SessionLocal()
    try:
        leads   = session.query(Lead).order_by(Lead.created_at.desc()).limit(200).all()
        updates = session.query(LeadUpdate).order_by(LeadUpdate.updated_at.desc()).limit(200).all()

        # Build lead lookup for company name on updates
        lead_map = {l.id: l for l in leads}

        events = []
        for l in leads:
            events.append({
                "type": "lead", "lead_id": l.id,
                "company_name": l.company_name, "client_name": l.client_name,
                "site_status": l.site_status, "stage": l.stage,
                "material": l.material, "quantity": l.quantity, "remarks": l.remarks,
                "sales_exec_name": l.sales_exec_name,
                "created_at": l.created_at.isoformat() if l.created_at else None,
            })
        for u in updates:
            parent = lead_map.get(u.lead_id)
            events.append({
                "type": "update", "lead_id": u.lead_id,
                "company_name": parent.company_name if parent else "—",
                "client_name":  parent.client_name  if parent else "—",
                "site_status": u.site_status, "stage": u.stage,
                "material": u.material, "quantity": u.quantity, "remarks": u.remarks,
                "sales_exec_name": u.sales_exec_name,
                "created_at": u.updated_at.isoformat() if u.updated_at else None,
            })

        # Sort all events newest first
        events.sort(key=lambda e: e["created_at"] or "", reverse=True)
        return jsonify({"events": events[:300]})
    finally:
        session.close()

@app.route("/admin/api/execs")
@require_admin
def admin_execs():
    """Per-exec performance summary."""
    from sqlalchemy import func
    session = SessionLocal()
    try:
        rows = (session.query(
                    Lead.sales_exec_id,
                    Lead.sales_exec_name,
                    func.count(Lead.id).label("total"),
                    func.max(Lead.created_at).label("last_active")
                )
                .group_by(Lead.sales_exec_id, Lead.sales_exec_name)
                .order_by(func.count(Lead.id).desc())
                .all())

        result = []
        for r in rows:
            q = session.query(Lead).filter(Lead.sales_exec_id == r.sales_exec_id)
            won  = q.filter(Lead.site_status == "Won").count()
            lost = q.filter(Lead.site_status == "Lost").count()
            result.append({
                "exec_id":     r.sales_exec_id,
                "name":        r.sales_exec_name or f"Exec #{r.sales_exec_id}",
                "total":       r.total,
                "won":         won,
                "lost":        lost,
                "in_progress": max(r.total - won - lost, 0),
                "last_active": r.last_active.isoformat() if r.last_active else None,
            })
        return jsonify({"execs": result})
    finally:
        session.close()


@app.route("/admin/api/execs/<int:exec_id>/target", methods=["GET", "POST"])
@require_admin
def admin_exec_target(exec_id):
    """GET or POST targets for a specific exec."""
    session = SessionLocal()
    try:
        target = session.query(ExecTarget).filter(
            ExecTarget.sales_exec_id == exec_id
        ).first()

        if request.method == "GET":
            if target:
                return jsonify({
                    "exec_id":       target.sales_exec_id,
                    "exec_name":     target.sales_exec_name,
                    "monthly_leads": target.monthly_leads,
                    "conversion_pct":target.conversion_pct,
                    "volume_m3":     target.volume_m3,
                })
            # Return defaults if no target set yet
            exec_name = session.query(Lead.sales_exec_name).filter(
                Lead.sales_exec_id == exec_id
            ).limit(1).scalar() or f"Exec #{exec_id}"
            return jsonify({
                "exec_id":        exec_id,
                "exec_name":      exec_name,
                "monthly_leads":  30,
                "conversion_pct": 40.0,
                "volume_m3":      500.0,
            })

        # POST — save/update target
        data = request.get_json(force=True) or {}
        exec_name = session.query(Lead.sales_exec_name).filter(
            Lead.sales_exec_id == exec_id
        ).limit(1).scalar() or f"Exec #{exec_id}"

        if not target:
            target = ExecTarget(sales_exec_id=exec_id, sales_exec_name=exec_name)
            session.add(target)

        target.sales_exec_name = exec_name
        if "monthly_leads"  in data: target.monthly_leads  = int(data["monthly_leads"])
        if "conversion_pct" in data: target.conversion_pct = float(data["conversion_pct"])
        if "volume_m3"      in data: target.volume_m3      = float(data["volume_m3"])
        session.commit()
        return jsonify({"status": "ok"})
    except Exception as e:
        session.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        session.close()


@app.route("/admin/api/execs/<int:exec_id>/progress")
@require_admin
def admin_exec_progress(exec_id):
    """
    Weekly and monthly progress for a specific exec vs their targets.
    Returns lead counts, status breakdown, and timeline data for charts.
    """
    from sqlalchemy import func
    from datetime import date, timezone as _tz

    session = SessionLocal()
    try:
        IST_OFF = timedelta(hours=5, minutes=30)

        now_ist   = datetime.now(_tz.utc) + IST_OFF
        # Month range
        month_start = now_ist.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        month_start_utc = month_start - IST_OFF
        # Week range (Mon–Sun)
        week_start = now_ist - timedelta(days=now_ist.weekday())
        week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start_utc = week_start - IST_OFF

        q_all   = session.query(Lead).filter(Lead.sales_exec_id == exec_id)
        q_month = q_all.filter(Lead.created_at >= month_start_utc)
        q_week  = q_all.filter(Lead.created_at >= week_start_utc)

        def status_breakdown(q):
            won  = q.filter(Lead.site_status == "Won").count()
            lost = q.filter(Lead.site_status == "Lost").count()
            neg  = q.filter(Lead.site_status == "Negotiation in Progress").count()
            vis  = q.filter(Lead.site_status == "Visited").count()
            quo  = q.filter(Lead.site_status == "Quoted").count()
            total = won + lost + neg + vis + quo
            return {"won": won, "lost": lost, "negotiating": neg,
                    "visited": vis, "quoted": quo, "total": total}

        monthly = status_breakdown(q_month)
        weekly  = status_breakdown(q_week)

        # Volume
        def total_volume(q):
            v = 0.0
            for l in q.filter(Lead.quantity.isnot(None)).all():
                try: v += float(l.quantity)
                except: pass
            return round(v, 2)

        monthly["volume"] = total_volume(q_month)
        weekly["volume"]  = total_volume(q_week)

        # Conversion rate
        monthly["conv_pct"] = round(monthly["won"] / monthly["total"] * 100, 1) if monthly["total"] else 0
        weekly["conv_pct"]  = round(weekly["won"]  / weekly["total"]  * 100, 1) if weekly["total"] else 0

        # Stage breakdown for this month (for pie chart)
        stage_rows = (
            session.query(Lead.stage, func.count(Lead.id))
            .filter(Lead.sales_exec_id == exec_id, Lead.created_at >= month_start_utc)
            .group_by(Lead.stage).all()
        )
        stage_breakdown = [{"stage": s or "—", "count": c} for s, c in stage_rows]

        # Daily lead count for past 30 days (for sparkline)
        thirty_ago = now_ist - timedelta(days=29)
        thirty_ago_utc = thirty_ago.replace(hour=0, minute=0, second=0, microsecond=0) - IST_OFF
        daily_rows = (
            session.query(Lead.created_at)
            .filter(Lead.sales_exec_id == exec_id, Lead.created_at >= thirty_ago_utc)
            .all()
        )
        # Bucket by IST date
        daily_map = {}
        for (dt,) in daily_rows:
            d = (dt + IST_OFF).strftime("%Y-%m-%d")
            daily_map[d] = daily_map.get(d, 0) + 1
        daily = [{"date": d, "count": daily_map.get(d, 0)}
                 for d in [(thirty_ago + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(30)]]

        # Target
        target = session.query(ExecTarget).filter(
            ExecTarget.sales_exec_id == exec_id
        ).first()
        targets = {
            "monthly_leads":  target.monthly_leads  if target else 30,
            "conversion_pct": target.conversion_pct if target else 40.0,
            "volume_m3":      target.volume_m3      if target else 500.0,
        }

        exec_name = session.query(Lead.sales_exec_name).filter(
            Lead.sales_exec_id == exec_id
        ).limit(1).scalar() or f"Exec #{exec_id}"

        return jsonify({
            "exec_id":        exec_id,
            "exec_name":      exec_name,
            "monthly":        monthly,
            "weekly":         weekly,
            "stage_breakdown":stage_breakdown,
            "daily_trend":    daily,
            "targets":        targets,
        })
    finally:
        session.close()


# ── Health check ──────────────────────────────────────────────────────────────
@app.route("/admin/api/export/csv")
@require_admin
def admin_export_csv():
    """Export all leads + their update history as a CSV file."""
    import csv, io
    from sqlalchemy import func

    session = SessionLocal()
    try:
        leads = session.query(Lead).order_by(Lead.created_at.asc()).all()
        lead_ids = [l.id for l in leads]

        # Latest update per lead
        latest = dict(
            session.query(LeadUpdate.lead_id, func.max(LeadUpdate.updated_at))
            .filter(LeadUpdate.lead_id.in_(lead_ids))
            .group_by(LeadUpdate.lead_id).all()
        ) if lead_ids else {}

        # Update counts
        counts = dict(
            session.query(LeadUpdate.lead_id, func.count(LeadUpdate.id))
            .filter(LeadUpdate.lead_id.in_(lead_ids))
            .group_by(LeadUpdate.lead_id).all()
        ) if lead_ids else {}

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "ID", "Company", "Contact Person", "Phone",
            "Work Status", "Stage", "Material", "Grade", "Quantity",
            "Remarks", "Location", "Sales Executive",
            "First Submitted (IST)", "Last Updated (IST)", "Total Updates"
        ])

        IST_OFFSET = timedelta(hours=5, minutes=30)

        def to_ist(dt):
            if not dt:
                return ""
            return (dt + IST_OFFSET).strftime("%d %b %Y %I:%M %p")

        for l in leads:
            last_upd = latest.get(l.id)
            writer.writerow([
                l.id, l.company_name, l.client_name, l.client_phone,
                l.site_status, l.stage, l.material, l.grade or "", l.quantity,
                l.remarks or "", l.location or "", l.sales_exec_name,
                to_ist(l.created_at),
                to_ist(last_upd) if last_upd else to_ist(l.created_at),
                counts.get(l.id, 0),
            ])

        output.seek(0)
        from flask import Response
        date_str = (datetime.now(timezone.utc) + timedelta(hours=5, minutes=30)).strftime("%Y-%m-%d")
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename=mcube_leads_{date_str}.csv"}
        )
    finally:
        session.close()


@app.route("/admin/api/export/excel")
@require_admin
def admin_export_excel():
    """Export all leads as a formatted .xlsx file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    from sqlalchemy import func
    import io

    session = SessionLocal()
    try:
        leads = session.query(Lead).order_by(Lead.created_at.asc()).all()
        lead_ids = [l.id for l in leads]

        latest = dict(
            session.query(LeadUpdate.lead_id, func.max(LeadUpdate.updated_at))
            .filter(LeadUpdate.lead_id.in_(lead_ids))
            .group_by(LeadUpdate.lead_id).all()
        ) if lead_ids else {}

        counts = dict(
            session.query(LeadUpdate.lead_id, func.count(LeadUpdate.id))
            .filter(LeadUpdate.lead_id.in_(lead_ids))
            .group_by(LeadUpdate.lead_id).all()
        ) if lead_ids else {}

        IST_OFFSET = timedelta(hours=5, minutes=30)

        def to_ist(dt):
            if not dt:
                return ""
            return (dt + IST_OFFSET).strftime("%d %b %Y %I:%M %p")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads"

        # ── Header row ─────────────────────────────────────────────────────────
        headers = [
            "ID", "Company", "Contact Person", "Phone",
            "Work Status", "Stage", "Material", "Grade", "Quantity",
            "Remarks", "Location", "Sales Executive",
            "First Submitted (IST)", "Last Updated (IST)", "Total Updates"
        ]
        col_widths = [6, 24, 20, 14, 20, 12, 16, 10, 12, 28, 22, 20, 22, 22, 14]

        header_fill   = PatternFill("solid", fgColor="1A237E")
        header_font   = Font(bold=True, color="FFFFFF", size=11)
        center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border   = Border(
            left=Side(style="thin", color="E8EAEF"),
            right=Side(style="thin", color="E8EAEF"),
            bottom=Side(style="thin", color="E8EAEF"),
        )

        # Status fill colours
        status_fills = {
            "Won":                    PatternFill("solid", fgColor="E8F5E9"),
            "Lost":                   PatternFill("solid", fgColor="FFEBEE"),
            "Negotiation in Progress":PatternFill("solid", fgColor="FFF3E0"),
            "Visited":                PatternFill("solid", fgColor="E8EAF6"),
            "Quoted":                 PatternFill("solid", fgColor="E8EAF6"),
        }

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center_align
            cell.border    = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 36
        ws.freeze_panes = "A2"  # freeze header row

        # ── Data rows ──────────────────────────────────────────────────────────
        alt_fill = PatternFill("solid", fgColor="F7F8FC")

        for row_idx, l in enumerate(leads, 2):
            last_upd = latest.get(l.id)
            row_data = [
                l.id, l.company_name, l.client_name, l.client_phone,
                l.site_status, l.stage, l.material, l.grade or "", l.quantity or "",
                l.remarks or "", l.location or "", l.sales_exec_name or "",
                to_ist(l.created_at),
                to_ist(last_upd) if last_upd else to_ist(l.created_at),
                counts.get(l.id, 0),
            ]

            row_fill = status_fills.get(l.site_status, PatternFill("solid", fgColor="FFFFFF"))
            if row_idx % 2 == 0 and l.site_status not in status_fills:
                row_fill = alt_fill

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill      = row_fill
                cell.border    = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=False)
            ws.row_dimensions[row_idx].height = 20

        # ── Summary sheet ──────────────────────────────────────────────────────
        ws2 = wb.create_sheet("Summary")
        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 14

        total = len(leads)
        won   = sum(1 for l in leads if l.site_status == "Won")
        lost  = sum(1 for l in leads if l.site_status == "Lost")
        date_str = to_ist(datetime.now(timezone.utc))

        summary_rows = [
            ("Mcube M3 — Lead Export", ""),
            (f"Generated at (IST)", date_str),
            ("", ""),
            ("Total Leads", total),
            ("Won", won),
            ("Lost", lost),
            ("In Progress", total - won - lost),
            ("Conversion Rate", f"{round(won/total*100, 1)}%" if total else "0%"),
        ]
        for r_idx, (label, value) in enumerate(summary_rows, 1):
            ws2.cell(row=r_idx, column=1, value=label).font = Font(bold=(r_idx in (1, 4, 5, 6, 7, 8)))
            ws2.cell(row=r_idx, column=2, value=value)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        from flask import Response
        date_label = (datetime.now(timezone.utc) + IST_OFFSET).strftime("%Y-%m-%d")
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=mcube_leads_{date_label}.xlsx"}
        )
    finally:
        session.close()


@app.route("/admin/api/test-digest", methods=["POST"])
@require_admin
def admin_test_digest():
    """Trigger the daily digest immediately — useful for testing."""
    from bot.daily_digest import send_daily_digests
    import os as _os
    try:
        send_daily_digests(bot=telegram_app.bot, loop=loop)
        owner = _os.getenv("DIGEST_OWNER_CHAT_ID", "").strip()
        return jsonify({
            "status": "ok",
            "recipients": "all execs + owner" if owner else "all execs only (DIGEST_OWNER_CHAT_ID not set in .env)"
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/admin/api/scheduler-status")
@require_admin
def admin_scheduler_status():
    """Inspect all scheduled jobs and their next fire times (IST)."""
    from datetime import timezone, timedelta
    IST = timezone(timedelta(hours=5, minutes=30))

    jobs = []
    for job in apscheduler.get_jobs():
        next_run = job.next_run_time
        if next_run:
            next_ist = next_run.astimezone(IST).strftime("%Y-%m-%d %H:%M:%S IST")
        else:
            next_ist = "not scheduled / paused"
        jobs.append({
            "id":       job.id,
            "name":     job.name,
            "next_run": next_ist,
            "trigger":  str(job.trigger),
        })

    now_ist = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S IST")
    print_scheduler_status()   # also print to server console
    return jsonify({
        "scheduler_running": apscheduler.running,
        "current_time_ist":  now_ist,
        "jobs":              jobs,
        "job_count":         len(jobs),
    })


# ── Health check ──────────────────────────────────────────────────────────────
@app.route("/health")
def health():
    return {"status": "running"}, 200

# ── Entrypoint ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    try:
        # use_reloader=False — reloader spawns a second process which double-inits the bot
        app.run(host="127.0.0.1", port=5001, use_reloader=False)
    except KeyboardInterrupt:
        async def shutdown():
            await telegram_app.stop()
            await telegram_app.shutdown()
        asyncio.run_coroutine_threadsafe(shutdown(), loop).result(timeout=10)