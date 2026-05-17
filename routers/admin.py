# routers/admin.py
"""
Tenant Admin API — secured by require_tenant_admin.

Base URL: /admin/...

Endpoints:
  GET  /admin/leads              — paginated lead list
  GET  /admin/leads/{id}         — single lead detail
  GET  /admin/leads/export       — Excel export
  GET  /admin/stats              — dashboard KPIs
  GET  /admin/executives         — distinct sales exec list
  POST /admin/config             — update tenant config key(s)
  GET  /admin/config             — read all config
  POST /admin/targets/{exec_id}  — set exec monthly targets
"""
import io
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, Query, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from dependencies import get_db, require_tenant_admin
from models.lead import Lead
from models.lead_update import LeadUpdate
from models.exec_target import ExecTarget
from models.tenant import Tenant
from models.tenant_config import TenantConfig, get_config, set_config

router = APIRouter(prefix="/admin", tags=["admin"])

_IST = timedelta(hours=5, minutes=30)


def _to_ist_str(dt: Optional[datetime]) -> Optional[str]:
    if not dt:
        return None
    aware = dt.replace(tzinfo=timezone.utc) if dt.tzinfo is None else dt
    return (aware + _IST).strftime("%d %b %Y, %I:%M %p")


# ── KPI stats ─────────────────────────────────────────────────────────────────

@router.get("/stats")
async def admin_stats(
    tenant: Tenant = Depends(require_tenant_admin),
    db: Session    = Depends(get_db),
):
    q     = db.query(Lead).filter(Lead.tenant_id == tenant.id)
    total = q.count()
    won   = q.filter(Lead.site_status == "Won").count()
    lost  = q.filter(Lead.site_status == "Lost").count()

    # Per-stage breakdown
    stages = db.query(Lead.stage, func.count(Lead.id)).filter(
        Lead.tenant_id == tenant.id
    ).group_by(Lead.stage).all()

    # Per-exec breakdown
    execs = db.query(
        Lead.sales_exec_name,
        func.count(Lead.id).label("total"),
    ).filter(Lead.tenant_id == tenant.id).group_by(Lead.sales_exec_name).all()

    # Monthly trend (last 6 months)
    now    = datetime.now(timezone.utc)
    months = []
    for i in range(5, -1, -1):
        month_start = (now.replace(day=1) - timedelta(days=30 * i)).replace(
            day=1, hour=0, minute=0, second=0, microsecond=0
        )
        month_end = (month_start + timedelta(days=32)).replace(day=1)
        count = q.filter(Lead.created_at >= month_start, Lead.created_at < month_end).count()
        months.append({"month": month_start.strftime("%b %Y"), "count": count})

    return {
        "total":       total,
        "won":         won,
        "lost":        lost,
        "in_progress": max(total - won - lost, 0),
        "stages":      [{"stage": s, "count": c} for s, c in stages],
        "executives":  [{"name": n,  "count": c} for n, c in execs],
        "trend":       months,
    }


# ── Leads list ─────────────────────────────────────────────────────────────────

@router.get("/leads")
async def admin_leads(
    page:           int = Query(1, ge=1),
    per_page:       int = Query(50, ge=1, le=200),
    exec_id:        Optional[int]  = None,
    stage:          Optional[str]  = None,
    status:         Optional[str]  = None,
    search:         Optional[str]  = None,
    date_from:      Optional[str]  = None,
    date_to:        Optional[str]  = None,
    tenant: Tenant  = Depends(require_tenant_admin),
    db: Session     = Depends(get_db),
):
    q = db.query(Lead).filter(Lead.tenant_id == tenant.id)

    if exec_id:  q = q.filter(Lead.sales_exec_id == exec_id)
    if stage:    q = q.filter(Lead.stage == stage)
    if status:   q = q.filter(Lead.site_status == status)
    if search:
        pat = f"%{search}%"
        q   = q.filter(
            Lead.company_name.ilike(pat) | Lead.client_name.ilike(pat) | Lead.client_phone.ilike(pat)
        )
    if date_from:
        q = q.filter(Lead.created_at >= datetime.fromisoformat(date_from))
    if date_to:
        q = q.filter(Lead.created_at <= datetime.fromisoformat(date_to))

    total = q.count()
    leads = q.order_by(Lead.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()

    return {
        "total": total,
        "page":  page,
        "pages": (total + per_page - 1) // per_page,
        "leads": [{
            "id":             l.id,
            "company_name":   l.company_name,
            "client_name":    l.client_name,
            "client_phone":   l.client_phone,
            "site_status":    l.site_status,
            "stage":          l.stage,
            "material":       l.material,
            "grade":          l.grade,
            "quantity":       l.quantity,
            "remarks":        l.remarks,
            "location":       l.location,
            "sales_exec_name": l.sales_exec_name,
            "sales_exec_id":  l.sales_exec_id,
            "created_at":     _to_ist_str(l.created_at),
            "last_followup_at": _to_ist_str(l.last_followup_at),
            "next_followup_date": _to_ist_str(l.next_followup_date),
        } for l in leads]
    }


@router.get("/leads/export")
async def export_leads(
    tenant: Tenant = Depends(require_tenant_admin),
    db: Session    = Depends(get_db),
):
    """Download all leads as an Excel file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    leads = db.query(Lead).filter(Lead.tenant_id == tenant.id).order_by(Lead.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = [
        "ID", "Company", "Contact", "Phone", "Status", "Stage",
        "Material", "Grade", "Qty (cum)", "Remarks", "Location",
        "Sales Exec", "Created (IST)", "Last Follow-up (IST)", "Next Follow-up (IST)",
    ]
    header_fill = PatternFill(fill_type="solid", fgColor="01696F")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, l in enumerate(leads, 2):
        ws.append([
            l.id, l.company_name, l.client_name, l.client_phone,
            l.site_status, l.stage, l.material, l.grade, l.quantity, l.remarks, l.location,
            l.sales_exec_name,
            _to_ist_str(l.created_at),
            _to_ist_str(l.last_followup_at),
            _to_ist_str(l.next_followup_date),
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    now_ist  = datetime.now(timezone(timedelta(hours=5, minutes=30)))
    filename = f"leads_{tenant.slug}_{now_ist.strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/leads/{lead_id}")
async def admin_lead_detail(
    lead_id: int,
    tenant: Tenant = Depends(require_tenant_admin),
    db: Session    = Depends(get_db),
):
    lead = db.query(Lead).filter(Lead.id == lead_id, Lead.tenant_id == tenant.id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    updates = (
        db.query(LeadUpdate)
        .filter(LeadUpdate.lead_id == lead_id)
        .order_by(LeadUpdate.updated_at.asc())
        .all()
    )
    return {
        "lead": {
            "id": lead.id, "company_name": lead.company_name, "client_name": lead.client_name,
            "client_phone": lead.client_phone, "site_status": lead.site_status,
            "stage": lead.stage, "material": lead.material, "grade": lead.grade,
            "quantity": lead.quantity, "remarks": lead.remarks, "location": lead.location,
            "latitude": lead.latitude, "longitude": lead.longitude,
            "sales_exec_name": lead.sales_exec_name,
            "created_at": _to_ist_str(lead.created_at),
        },
        "updates": [{
            "id": u.id, "stage": u.stage, "site_status": u.site_status,
            "quantity": u.quantity, "remarks": u.remarks,
            "updated_at": _to_ist_str(u.updated_at),
        } for u in updates]
    }


# ── Executives list ─────────────────────────────────────────────────────────────────

@router.get("/executives")
async def admin_executives(
    tenant: Tenant = Depends(require_tenant_admin),
    db: Session    = Depends(get_db),
):
    rows = db.query(
        Lead.sales_exec_id,
        Lead.sales_exec_name,
        func.count(Lead.id).label("total"),
    ).filter(Lead.tenant_id == tenant.id).group_by(
        Lead.sales_exec_id, Lead.sales_exec_name
    ).all()
    return [{"id": r.sales_exec_id, "name": r.sales_exec_name, "total": r.total} for r in rows]


# ── Tenant config ────────────────────────────────────────────────────────────────

@router.get("/config")
async def admin_get_config(
    tenant: Tenant = Depends(require_tenant_admin),
    db: Session    = Depends(get_db),
):
    rows = db.query(TenantConfig).filter(TenantConfig.tenant_id == tenant.id).all()
    import json
    return {r.key: json.loads(r.value_json) if r.value_json else None for r in rows}


@router.post("/config")
async def admin_set_config(
    request: Request,
    tenant: Tenant = Depends(require_tenant_admin),
    db: Session    = Depends(get_db),
):
    """Body: {"stages": [...], "materials": [...], "brand_name": "...", ...}"""
    data = await request.json()
    for key, value in data.items():
        set_config(db, tenant.id, key, value)
    return {"status": "ok", "updated": list(data.keys())}


# ── Exec targets ────────────────────────────────────────────────────────────────────

@router.post("/targets/{exec_id}")
async def admin_set_target(
    exec_id: int,
    request: Request,
    tenant: Tenant = Depends(require_tenant_admin),
    db: Session    = Depends(get_db),
):
    data   = await request.json()
    target = db.query(ExecTarget).filter(
        ExecTarget.sales_exec_id == exec_id,
        ExecTarget.tenant_id == tenant.id,
    ).first()
    if not target:
        target = ExecTarget(sales_exec_id=exec_id, tenant_id=tenant.id)
        db.add(target)

    if "monthly_leads"   in data: target.monthly_leads   = int(data["monthly_leads"])
    if "conversion_pct"  in data: target.conversion_pct  = float(data["conversion_pct"])
    if "volume_m3"       in data: target.volume_m3        = float(data["volume_m3"])

    db.commit()
    return {"status": "ok"}
